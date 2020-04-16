import os
import csv
import openpyxl

for file in os.listdir('.'):
    if file.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file)

        for sheetName in wb.get_sheet_names():
            sheet = wb.get_sheet_by_name(sheetName)
            nameSplice =  file[:-5]
            csvFile = open(nameSplice + '_' + sheetName + '.csv' , 'w', newline='')
            csvWriter = csv.writer(csvFile)

            for rowNum in range(1, sheet.max_row + 1):
                rowData = []
                for colNum in range(1, sheet.max_column + 1):
                    rowData.append(sheet.cell(row = rowNum, column = colNum).value)

                for row in rowData:
                    csvWriter.writerow(row)
            csvFile.close()