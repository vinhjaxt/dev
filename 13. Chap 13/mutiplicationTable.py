import sys
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

number = int(sys.argv[1])
cells = number + 1

wb = openpyxl.Workbook()
sheet = wb.active

makeBold =  Font(bold=True)

while cells > 1 :
    sheet.cell(row = cells, column = 1).value = cells - 1
    sheet.cell(row = 1, column = cells).value = cells - 1
    sheet.cell(row = cells, column = 1).font = makeBold
    sheet.cell(row = 1, column = cells).font = makeBold

    cells -= 1

colLength = number + 1
count = 0
while count < number:
    colLetter = get_column_letter(sheet.max_column - count)
    while colLength > 1 :
        sheet[colLetter + str(colLength)] = ('=SUM(' + colLetter + '1*A' + str(colLength) + ')')
        colLength -= 1
    colLength = number + 1
    count += 1

wb.save('multiTable.xlsx')
                                               